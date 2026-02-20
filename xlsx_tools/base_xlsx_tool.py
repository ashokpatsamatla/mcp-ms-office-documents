import io
import logging
import re
from typing import List
from openpyxl import Workbook

from upload_tools import upload_file
from .helpers import (
    parse_table,
    add_table_to_sheet,
)

logger = logging.getLogger(__name__)

# Pattern for multi-sheet heading: ## Sheet: Name
SHEET_HEADING_PATTERN = re.compile(r'^##\s+Sheet:\s+(.+)$')


def markdown_to_excel(markdown_content: str) -> str:
    """Convert Markdown to Excel workbook (focused on tables and headers).

    Always starts from an empty Workbook (no templates).
    Supports multiple sheets via '## Sheet: Name' headings.
    """
    logger.info("Starting markdown_to_excel conversion")

    # Create a fresh workbook (no templates)
    wb = Workbook()
    ws = wb.active

    # Set default worksheet title
    try:
        ws.title = "Data Report"
    except Exception:
        # Some Excel title errors may occur if title invalid; ignore and keep default
        logger.debug("Could not set worksheet title; keeping default")

    # Split content into lines
    lines: List[str] = markdown_content.split('\n')

    # Counters for a short summary
    headers_count = 0
    tables_count = 0

    # Per-sheet state
    current_row = 1
    table_counter = 1
    table_positions = {}  # Track where each table starts
    first_sheet_named = False  # Whether we've set a name for the first sheet
    i = 0

    try:
        while i < len(lines):
            line = lines[i].strip()

            # Skip empty lines
            if not line:
                i += 1
                continue

            # Check for sheet heading: ## Sheet: Name
            sheet_match = SHEET_HEADING_PATTERN.match(line)
            if sheet_match:
                sheet_name = sheet_match.group(1).strip()
                if not first_sheet_named and current_row == 1:
                    # Rename the default sheet instead of creating a new one
                    try:
                        ws.title = sheet_name
                    except Exception:
                        logger.debug("Could not rename worksheet to '%s'", sheet_name)
                else:
                    # Create a new worksheet
                    ws = wb.create_sheet(title=sheet_name)
                    current_row = 1
                    table_counter = 1
                    table_positions = {}
                first_sheet_named = True
                i += 1
                continue

            # Headers
            if line.startswith('#'):
                header_level = len(line) - len(line.lstrip('#'))
                header_text = line.lstrip('#').strip()

                cell = ws.cell(row=current_row, column=1)
                cell.value = header_text

                # Style headers based on level
                from openpyxl.styles import Font  # local import to keep top clean
                if header_level == 1:
                    cell.font = Font(size=16, bold=True, color="2F5597")
                elif header_level == 2:
                    cell.font = Font(size=14, bold=True, color="4472C4")
                else:
                    cell.font = Font(size=12, bold=True)

                headers_count += 1
                logger.debug("Header (level %d): %s", header_level, header_text)

                current_row += 2  # Add space after headers
                i += 1

            # Tables
            elif line.startswith('|'):
                table_data, i = parse_table(lines, i)
                if table_data:
                    # Record this table's position
                    table_key = f"T{table_counter}"
                    table_positions[table_key] = current_row

                    # Process the table
                    start_row_before = current_row
                    current_row = add_table_to_sheet(table_data, ws, current_row, table_positions)
                    _row_count = current_row - start_row_before - 2  # subtract header and spacing

                    tables_count += 1
                    logger.debug("Added table #%d with %d rows", tables_count, len(table_data))
                    table_counter += 1

            # Skip other content
            else:
                i += 1

    except Exception as e:
        logger.error("Error generating Excel workbook: %s", str(e), exc_info=True)
        raise RuntimeError(f"Error generating Excel workbook: {e}") from e

    # Save workbook to BytesIO and upload via existing helper
    file_object = io.BytesIO()
    try:
        logger.info("Saving Excel workbook to memory buffer")
        wb.save(file_object)
        file_object.seek(0)
        result = upload_file(file_object, "xlsx")
        logger.info("Excel upload completed (headers=%d, tables=%d)", headers_count, tables_count)
        return result
    except Exception as e:
        logger.error("Error saving/uploading Excel workbook: %s", str(e), exc_info=True)
        raise RuntimeError(f"Error saving/uploading Excel workbook: {e}") from e
    finally:
        file_object.close()
