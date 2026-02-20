"""Tests for Excel (xlsx) multi-sheet support.

These tests verify that the markdown to Excel conversion handles
the '## Sheet: Name' heading syntax correctly for multi-sheet workbooks.
"""

import sys
from pathlib import Path
from unittest.mock import patch, MagicMock

# Add project root to path for imports
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

import pytest
from openpyxl import Workbook, load_workbook
import io

# We test the internal parsing logic, mocking the upload step.
from xlsx_tools.base_xlsx_tool import markdown_to_excel


def _create_workbook_from_markdown(markdown_content: str) -> Workbook:
    """Helper that runs markdown_to_excel but intercepts the workbook before upload.

    Patches upload_file to capture the BytesIO and returns a loaded Workbook.
    """
    captured = {}

    def fake_upload(file_obj, suffix):
        captured['data'] = file_obj.read()
        file_obj.seek(0)
        return "https://fake-url/test.xlsx"

    with patch("xlsx_tools.base_xlsx_tool.upload_file", side_effect=fake_upload):
        markdown_to_excel(markdown_content)

    wb = load_workbook(io.BytesIO(captured['data']))
    return wb


# Output directory for test files
OUTPUT_DIR = Path(__file__).parent / "output" / "xlsx"


@pytest.fixture(scope="module", autouse=True)
def setup_output_dir():
    """Create output directory if it doesn't exist."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    yield


class TestMultiSheet:
    """Tests for multi-sheet Excel workbooks via ## Sheet: Name."""

    def test_single_sheet_default_name(self):
        """Markdown without ## Sheet: heading → single sheet named 'Data Report'."""
        markdown = """# Report

| Name | Value |
|------|-------|
| A    | 1     |
"""
        wb = _create_workbook_from_markdown(markdown)
        assert len(wb.sheetnames) == 1
        assert wb.sheetnames[0] == "Data Report"

    def test_two_sheets(self):
        """Markdown with two ## Sheet: headings → two sheets."""
        markdown = """## Sheet: Revenue

| Quarter | Amount |
|---------|--------|
| Q1      | 1000   |
| Q2      | 1200   |

## Sheet: Expenses

| Quarter | Amount |
|---------|--------|
| Q1      | 800    |
| Q2      | 900    |
"""
        wb = _create_workbook_from_markdown(markdown)
        assert len(wb.sheetnames) == 2
        assert "Revenue" in wb.sheetnames
        assert "Expenses" in wb.sheetnames

    def test_sheet_names_correct(self):
        """Verify sheet names are correctly set from headings."""
        markdown = """## Sheet: Summary

| Metric | Value |
|--------|-------|
| Total  | 100   |

## Sheet: Detail Data

| Item | Count |
|------|-------|
| A    | 50    |
| B    | 50    |
"""
        wb = _create_workbook_from_markdown(markdown)
        assert wb.sheetnames[0] == "Summary"
        assert wb.sheetnames[1] == "Detail Data"

    def test_data_on_correct_sheets(self):
        """Verify tables land on the correct sheets."""
        markdown = """## Sheet: Sheet1

| Col1 | Col2 |
|------|------|
| X    | Y    |

## Sheet: Sheet2

| ColA | ColB |
|------|------|
| M    | N    |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws1 = wb["Sheet1"]
        ws2 = wb["Sheet2"]

        # Check data exists on Sheet1 (header row + data row)
        # The exact row depends on spacing; just verify the values exist somewhere
        sheet1_values = []
        for row in ws1.iter_rows(values_only=True):
            sheet1_values.extend([v for v in row if v is not None])
        assert "X" in sheet1_values or "x" in str(sheet1_values).lower()

        sheet2_values = []
        for row in ws2.iter_rows(values_only=True):
            sheet2_values.extend([v for v in row if v is not None])
        assert "M" in sheet2_values or "m" in str(sheet2_values).lower()

    def test_three_sheets(self):
        """Test creating three sheets."""
        markdown = """## Sheet: Alpha

| A |
|---|
| 1 |

## Sheet: Beta

| B |
|---|
| 2 |

## Sheet: Gamma

| C |
|---|
| 3 |
"""
        wb = _create_workbook_from_markdown(markdown)
        assert len(wb.sheetnames) == 3
        assert wb.sheetnames == ["Alpha", "Beta", "Gamma"]

    def test_backwards_compatible_no_sheet_heading(self):
        """Without any ## Sheet: headings, everything goes to 'Data Report'."""
        markdown = """# My Report

| Name | Age |
|------|-----|
| Alice | 30 |
| Bob   | 25 |
"""
        wb = _create_workbook_from_markdown(markdown)
        assert len(wb.sheetnames) == 1
        assert wb.sheetnames[0] == "Data Report"


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])

